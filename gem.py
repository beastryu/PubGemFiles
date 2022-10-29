'''Seth Harrison 10/19/2022
Must ensure that python(obviously), openpyxl, and spicy are installed on your system'''

import sys
import os
import random
from scipy import stats
from collections import Counter
import collections
import openpyxl

#Defining some varibles with generic data


DC = 0
searchType = "gemstones"
hoursSpent = 5
numberFound = 10
priceTotal = 0
pref = "Duvals"
outputArray = []










#Asking for the user inputs

while 1 == 1:
	userInput = input("Were you looking for Hardstones or Gemstones?").lower()

	if userInput == "gemstones" or userInput == "gemstone": 
		searchType = "gemstones"
		break
	elif userInput == "hardstones" or userInput == "hardstone":
		searchType = "hardstones"
		break
	elif userInput != "gemstones" or userInput != "hardstones":
		print("Im sorry. You need to enter a vaild gemtype.")

hoursSpent = int(input("How many hours did you spend?"))



if searchType == "gemstones":
	
	class GemClass:
		def __init__(self, rarity, mRow, mColumn):
			
			self.rarity = rarity
			self.mRow = mRow
			self.mColumn = mColumn
			self.count = 0

	common = 0
	uncommon = 0
	rare = 0
	veryRare = 0
	legendary = 0
	
	
	print("Underground: -5 \nMountains: -4 \nRiver: -3 \nMountainous Region: -2\nDesert: -2\nJungle: -2\nCoastal Area: -1\nForest: +1\nArctic Region: +2 \nInside city walls: +3 \nInside most buildings: +5")
	DC = input("What is the DC modifier?(Subtract your proficiency bonus here, the base is 15.)")
	pref = input("And whose gems preference list are we using?")
	
	prefList = (pref + "PreferenceList.xlsx")
	wb_prefList = openpyxl.load_workbook(prefList)
	sheet_prefList = wb_prefList.active
	
	commonRow = 20
	uncommonRow = 10
	rareRow = 6
	
	veryRareRow = 10
	legendaryRow = 9

	commonColumn = 1
	uncommonColumn = 2
	rareColumn = 3
	veryRareColumn = 4 
	legendaryColumn = 5

	commonArray = []
	uncommonArray = []
	rareArray = []

	#Checks to see how many gems are successfully found

	for i in range(hoursSpent):
		if int(DC) < random.randint(1, 20):
			numberFound += 1
	print(numberFound)



	for i in range(commonRow):
		prefCell = sheet_prefList.cell((2+i), column = 1)
		commonArray.append(prefCell.value)

	for i in range(uncommonRow):
		prefCell = sheet_prefList.cell((2+i), column = 2)
		commonArray.append(prefCell.value)

	for i in range(rareRow):
		prefCell = sheet_prefList.cell((2+i), column = 3)
		commonArray.append(prefCell.value)
 

	#For each found determine the rarity	

	for i in range(numberFound):
		rarity = random.randint(1, 20)

		if rarity <= 6: 
			common += 1
		elif rarity >= 7 and rarity <= 11:
			uncommon += 1
		elif rarity >= 12 and rarity <= 15:
			rare += 1
		elif rarity >= 16 and rarity <= 18:
			veryRare += 1
		elif rarity >= 19:
			legendary += 1
		else:
			print(ERROR)


	
	print("You found " + str(common) + " Common gems\n" + str(uncommon) + " Uncommon gems\n" + str(rare) + " Rare gems\n" + str(veryRare) + " Very Rare\n" + str(legendary) + " Legendary gems")
	print(commonArray)


elif searchType == "hardstones":
	hardstoneList = "HardstoneList.xlsx"
	wb_hardstoneList = openpyxl.load_workbook(hardstoneList)
	sheet_hardstoneList = wb_hardstoneList.active
	row = sheet_hardstoneList.max_row
	
	for i in range(hoursSpent):
		
		stoneRow = random.randint(2, row)
		stoneCell = sheet_hardstoneList.cell(stoneRow, column = 1)
		priceCell = sheet_hardstoneList.cell(stoneRow, column = 2)
		numberOfStones = (random.randint(5, 60) - 4)
		priceTotal += (numberOfStones * priceCell.value)
		print(str(numberOfStones) + " " + stoneCell.value + " worth " + str((numberOfStones) * priceCell.value) + "Gp")
	else:
		print(str(priceTotal) + "Gp Total")

#Testing Data
print("searchType " + searchType)
print("hoursSpent " + str(hoursSpent))
print("DC " + str(DC))
print("numberFound " + str(numberFound))





# Give the location of the file 
gemsList = "GemList.xlsx"
  
# To open the workbook 
# workbook object is created 
wb_gemsList = openpyxl.load_workbook(gemsList) 
  
# Get workbook active sheet object 
# from the active attribute 
sheet_gemsList = wb_gemsList.active 
  
# Cell objects also have a row, column, 
# and coordinate attributes that provide 
# location information for the cell. 
  
# Note: The first row or 
# column integer is 1, not 0. 
  
# Cell object is created by using 
# sheet object's cell() method. 
cell_gemsList = sheet_gemsList.cell(row = 1, column = 1) 
  
# Print value of cell object 
# using the value attribute 
print(cell_gemsList.value) 

