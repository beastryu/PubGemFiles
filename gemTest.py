'''Seth Harrison 10/19/2022
Must ensure that python(obviously), and spicy are installed on your system'''

import sys
import os
import openpyxl 
import random
from scipy import stats

#Defining some varibles with generic data


DC = 0
searchType = "gemstones"
hoursSpent = 5
numberFound = 0
priceTotal = 0
pref = "Duvals"

class GemClass:
		def __init__(self, rarity, mRow, column):
			
			self.rarity = rarity
			self.mRow = mRow
			self.column = column
			self.count = 0
			self.list = []

		def countUp(self):
			self.count += 1




'''
#Asking for the user inputs

while 1 == 1:
	userInput = input("Were you looking for Hardstones or Gemstones?").lower()

	if userInput == "gemstones" or userInput == "gemstone": 
		searchType = "gemstones"
		break
	elif userInput == "hardstones" or userInput == "hardstone":
		searchType = "hardstones"
		break
	elif userInput != "gemstones" or userInput != "hardstones":
		print("Im sorry. You need to enter a vaild gemtype.")

hoursSpent = int(input("How many hours did you spend?"))
'''


if searchType == "gemstones":
	
	
	
	# w is the list, mu is the most likely index, sd is the tightness of the selection

	def weighted_choices(w, mu, sd, *, k=1):
		weights = stats.norm(mu, sd).pdf(range(len(w)))
		return choices(w, weights=weights, k=k)	

	CommonClass = GemClass("common", 20, 1)
	UncommonClass = GemClass("uncommon", 10, 2)
	RareClass = GemClass("uncommon", 6, 3)
	

	VeryRareClass = GemClass("uncommon", 10, 4)
	LegendaryClass = GemClass("uncommon", 9, 5)
	
	'''
	print("Underground: -5 \nMountains: -4 \nRiver: -3 \nMountainous Region: -2\nDesert: -2\nJungle: -2\nCoastal Area: -1\nForest: +1\nArctic Region: +2 \nInside city walls: +3 \nInside most buildings: +5")
	DC = input("What is the DC modifier?(Subtract your proficiency bonus here, the base is 15.)")
	pref = input("And whose gems preference list are we using?")
	'''

	prefSheet = (pref + "PreferenceList.xlsx")
	wb_prefSheet = openpyxl.load_workbook(prefSheet)
	sheet_prefSheet = wb_prefSheet.active

	fullSheet = ("GemList.xlsx")
	wb_fullSheet = openpyxl.load_workbook(fullSheet)
	sheet_fullSheet = wb_fullSheet.active
	#Checks to see how many gems are successfully found

	for i in range(hoursSpent):
		if int(DC) < random.randint(1, 20):
			numberFound += 1
	print(numberFound)



	for i in range(CommonClass.mRow):
		prefCell = sheet_prefSheet.cell((2+i), column = CommonClass.column)
		CommonClass.list.append(prefCell.value)

	for i in range(UncommonClass.mRow):
		prefCell = sheet_prefSheet.cell((2+i), column = UncommonClass.column)
		UncommonClass.list.append(prefCell.value)
	
	for i in range(RareClass.mRow):
		prefCell = sheet_prefSheet.cell((2+i), column = RareClass.column)
		RareClass.list.append(prefCell.value) 

	print(CommonClass.list)

 

	#For each found determine the rarity	

	for i in range(numberFound):
		rarity = random.randint(1, 20)

		if rarity <= 6: 
			CommonClass.countUp()
		elif rarity >= 7 and rarity <= 11:
			UncommonClass.countUp()
		elif rarity >= 12 and rarity <= 15:
			RareClass.countUp()
		elif rarity >= 16 and rarity <= 18:
			VeryRareClass.countUp()
		elif rarity >= 19:
			LegendaryClass.countUp()
		else:
			print(ERROR)


	
	print("You found " + str(CommonClass.count) + " Common gems\n" + str(UncommonClass.count) + " Uncommon gems\n" + str(RareClass.count) + " Rare gems\n" + str(VeryRareClass.count) + " Very Rare\n" + str(LegendaryClass.count) + " Legendary gems")
	print(CommonClass.list)