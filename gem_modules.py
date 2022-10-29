import sys
import os
import random
from scipy import stats
from collections import Counter
import collections
import openpyxl

fullSheet = ("GemList.xlsx")
wb_fullSheet = openpyxl.load_workbook(fullSheet)
gemSheet = wb_fullSheet.active

class GemClass():

    allGemValues = {
        'Common': [1, 50],
        'Uncommon': [2, 44],
        'Rare': [3, 18],
        'VeryRare': [4, 10],
        'Legendary': [5, 9]

    }

    prefSheetValues = {
        'Common': [1, 20],
        'Uncommon': [2, 10],
        'Rare': [3, 6],
    }

    allGemWeights = {
        'Common': 8.9,
        'Uncommon': 1,
        'Rare': 1,
        'VeryRare': 1,
        'Legendary': 1
    }

    RarityLists = {
        'Common': [],
        'Uncommon': [],
        'Rare': [],
        'VeryRare': [],
        'Legendary': []
    }
    '''
        'commonColumn': 1
        'commonRow': 50
        'uncommonColumn': 2
        'uncommonRow': 44
        'rareColumn': 3
        'rareRow ': 18
        'veryRareColumn': 4
        'veryRareRow': 10
        'legendaryColumn': 5
        'legendaryRow': 9

        common
        uncommon
        rare
        veryRare
        legendary
    '''
            
class gemRarityClass(GemClass):
        def __init__(self, rarity):
            
            self.rarity = rarity
            self.pref = ""
            self.count = 0
            self.midPoint = 0
            self.list = []
            self.deque = collections.deque()
            self.outputArray = []

        def countUp(self):
            self.count += 1




# w is the list, mu is the most likely index, sd is the tightness of the selection
def getWeightedChoices(w, mu, sd, *, k=1):
        weights = stats.norm(mu, sd).pdf(range(len(w)))
        return random.choices(w, weights=weights, k=k)

def popRarityLists(obj):
    obj = obj
    for (k, v) in obj.allGemValues.items():
        rarity = k
        obj.RarityLists.update({k:[]})
        column = obj.allGemValues.get(k)[0]
        rows = obj.allGemValues.get(k)[1]


        for i in range(rows):
            gemCell = gemSheet.cell((1+i), column)
            obj.RarityLists.get(rarity).append(gemCell.value)

#Create an array with pref values

def popPrefList(c, pref):
        
        pref = pref
        for i in range(c.prefSheetValues.get(c.rarity)[1]):
            prefCell = pref.cell((1+i), column = c.prefSheetValues.get(c.rarity)[0])
            if pref is str(): c.list.append(prefCell.value)


#Make a DeQue from the prev Array
def popDeque(c):
    for i in c.list:
        if (c.list.index(i) % 2 == 0):
            c.deque.append(i)
        else:
            c.deque.appendleft(i)
    c.midPoint = (1 + (len(c.deque))) / 2

#Add all unprefered gems to the deque

def combineList(c):
    tempCounter = Counter(c.RarityLists.get((c.rarity)))
    pref = Counter(c.list)
    tempCounter.subtract(pref)
    tempCounter += collections.Counter()
    tempArray = []
    tempArray = list(tempCounter)
    
    for i in tempArray:
        if (tempArray.index(i) % 2 == 0):
            c.deque.append(i)
        else:
            c.deque.appendleft(i)
    c.midPoint = (1 + (len(c.deque))) / 2

def popClass(c, pref):
    popRarityLists(c)
    if c.rarity in c.prefSheetValues.keys(): popPrefList(c, pref)
    popDeque(c)
    combineList(c)

def result(c):
    w = c.allGemWeights.get(c.rarity)
    c.outputArray = getWeightedChoices(c.deque, c.midPoint, w, k=c.count)
    for i in range(len(c.outputArray)):
        size = random.randint(1, 3)
        if size == 1:
            c.outputArray[i] = ("Small " +c.outputArray[i])
        elif size == 2:
            c.outputArray[i] = ("Medium " +c.outputArray[i])
        elif size == 3:
            c.outputArray[i] = ("Large " +c.outputArray[i])
    return (c.outputArray)